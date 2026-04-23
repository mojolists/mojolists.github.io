#!/usr/bin/env python3
"""
Build _data/vinyl.json from the master vinyl spreadsheet.

Reads: ../../../NFT Creator Music/vinyl-master.xlsx (or --in override)
Writes: ../_data/vinyl.json

The JSON is consumed by /vinyl/ on mojolists.com.
"""

from __future__ import annotations

import argparse
import datetime
import json
import re
from pathlib import Path

from openpyxl import load_workbook

# ---------- defaults --------------------------------------------------------

DEFAULT_XLSX = Path(__file__).resolve().parent.parent.parent / "NFT Creator Music" / "vinyl-master.xlsx"
DEFAULT_OUT = Path(__file__).resolve().parent.parent / "_data" / "vinyl.json"
DEFAULT_NFC_DIR = Path(__file__).resolve().parent.parent / "NFC"

# ---------- helpers ---------------------------------------------------------

_slug_non_alnum = re.compile(r"[^a-z0-9]+")


def slugify(*parts):
    s = "-".join(str(p or "").lower() for p in parts if p)
    s = _slug_non_alnum.sub("-", s).strip("-")
    return s


def clean(v):
    if v is None:
        return ""
    return str(v).strip()


def parse_genre(v):
    if not v:
        return []
    parts = re.split(r"[,;/|]", str(v))
    return [p.strip() for p in parts if p.strip()]


def normalize_format(fmt_raw, vinyl_type):
    vt = (vinyl_type or "").strip()
    if vt == "LP":
        return "LP"
    if vt in ('12"', "12in", "12 inch", "12"):
        return '12"'
    if vt in ('10"', "10in", "10 inch", "10"):
        return '10"'
    fl = (fmt_raw or "").lower()
    if "10" in fl:
        return '10"'
    if "12" in fl:
        return '12"'
    return "LP"


# ---------- main ------------------------------------------------------------

def build(xlsx_path, out_path, nfc_dir):
    wb = load_workbook(xlsx_path, data_only=True)
    sheet_name = next((n for n in wb.sheetnames if n != "Summary"), wb.sheetnames[0])
    ws = wb[sheet_name]

    header = [clean(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(header)}

    required = ["Artist", "Title", "VinylType", "CollectionFolder"]
    missing = [c for c in required if c not in col]
    if missing:
        raise SystemExit("ERROR: spreadsheet missing required columns: " + str(missing))

    existing_shorts = set()
    if nfc_dir.exists():
        for p in nfc_dir.glob("*.md"):
            if p.name.lower() == "readme.md":
                continue
            existing_shorts.add(p.stem.lower())

    records = []
    seen_ids = set()
    locations = {}
    formats = {}
    genres = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        def g(name):
            i = col.get(name)
            return clean(row[i]) if i is not None and i < len(row) else ""

        artist = g("Artist")
        album = g("Title")
        if not artist and not album:
            continue

        vt_raw = g("VinylType")
        fmt_raw = g("Format (raw)") or g("Format")
        fmt = normalize_format(fmt_raw, vt_raw)

        slug = slugify(artist, album) or ("row-" + str(row_idx))
        release_id = g("Discogs release_id")
        rec_id = release_id or slug

        if rec_id in seen_ids:
            continue
        seen_ids.add(rec_id)

        genre_list = parse_genre(g("Genre")) or parse_genre(g("Genres"))

        try:
            rating_val = int(float(g("Rating"))) if g("Rating") else None
            if rating_val == 0:
                rating_val = None
        except ValueError:
            rating_val = None

        location = g("CollectionFolder") or "Unsorted"
        has_shorts = slug.lower() in existing_shorts
        shorts_url = ("/NFC/" + slug + "/") if has_shorts else ""

        rec = {
            "id": rec_id,
            "slug": slug,
            "artist": artist,
            "album": album,
            "year": g("Released"),
            "label": g("Label"),
            "format": fmt,
            "location": location,
            "genre": genre_list,
            "rating": rating_val,
            "discogs_url": g("Discogs URL"),
            "wiki_url": g("Wikipedia URL"),
            "mojolists_review_url": g("Mojolists review URL"),
            "shorts_url": shorts_url,
            "has_shorts": has_shorts,
            "cover": "covers/" + slug + ".jpg",
            "chart_peak": g("Chart peak (Billboard 200)") or g("Chart peak"),
            "key_tracks": g("Key tracks"),
            "catchy_fact": g("Catchy fact"),
        }
        records.append(rec)

        locations[location] = locations.get(location, 0) + 1
        formats[fmt] = formats.get(fmt, 0) + 1
        for g_ in genre_list:
            genres[g_] = genres.get(g_, 0) + 1

    records.sort(key=lambda r: (r["artist"].lower(), r["album"].lower()))

    payload = {
        "generated_at": datetime.datetime.utcnow().isoformat() + "Z",
        "count": len(records),
        "records": records,
        "facets": {
            "locations": sorted(locations.items(), key=lambda x: -x[1]),
            "formats": sorted(formats.items(), key=lambda x: -x[1]),
            "genres": sorted(genres.items(), key=lambda x: -x[1]),
            "shorts_done": sum(1 for r in records if r["has_shorts"]),
            "shorts_pending": sum(1 for r in records if not r["has_shorts"]),
        },
    }

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    return payload


def main():
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--in", dest="xlsx", type=Path, default=DEFAULT_XLSX)
    p.add_argument("--out", dest="out", type=Path, default=DEFAULT_OUT)
    p.add_argument("--nfc", dest="nfc", type=Path, default=DEFAULT_NFC_DIR)
    args = p.parse_args()

    if not args.xlsx.exists():
        raise SystemExit("ERROR: xlsx not found: " + str(args.xlsx))

    payload = build(args.xlsx, args.out, args.nfc)

    print("Wrote " + str(args.out))
    print("  records: " + str(payload["count"]))
    print("  formats: " + str(dict(payload["facets"]["formats"])))
    print("  locations: " + str(dict(payload["facets"]["locations"])))
    print("  shorts done/pending: " + str(payload["facets"]["shorts_done"]) + " / " + str(payload["facets"]["shorts_pending"]))


if __name__ == "__main__":
    main()
